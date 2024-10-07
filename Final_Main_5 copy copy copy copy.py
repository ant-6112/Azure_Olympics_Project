import sys
import pandas as pd
import xlsxwriter
from PyQt5.QtWidgets import (QApplication, QMainWindow, QFileDialog, QWidget, QVBoxLayout,
                             QHBoxLayout, QListWidget, QPushButton, QLabel, QTreeWidget,
                             QTreeWidgetItem, QComboBox, QTabWidget, QCheckBox, QScrollArea,
                             QInputDialog, QLineEdit, QTextEdit, QSplitter, QFrame, QHeaderView,
                             QTabWidget, QToolButton, QTableWidget, QTableWidgetItem, QMessageBox)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QFont
from jinja2 import Template, Environment, meta
import csv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os

class VarianceCommentaryTool(QMainWindow):
    def __init__(self):
        super().__init__()
        self.df1 = None
        self.df2 = None
        self.commentary_data = {}
        self.tab_data = []
        self.df1_paths = {}
        self.df2_paths = {}
        self.configurations = []
        self.current_config_index = 0
        self.initUI()

    def initUI(self):
            self.setWindowTitle('Variance Commentary Tool')
            self.setGeometry(100, 100, 1600, 900)

            #Central Widget
            central_widget = QWidget()
            self.setCentralWidget(central_widget)
            main_layout = QHBoxLayout(central_widget)

            #Left Widget
            self.left_widget = QWidget()
            self.left_widget.setFixedWidth(50)
            left_layout = QVBoxLayout(self.left_widget)
            left_layout.setContentsMargins(0, 0, 0, 0)

            #Toggle Button
            self.toggle_nav_button = QPushButton('◀', self)
            self.toggle_nav_button.setFixedSize(50, 100)
            self.toggle_nav_button.clicked.connect(self.toggle_navbar)
            left_layout.addWidget(self.toggle_nav_button)

            #Tab Widget
            self.tab_widget = QTabWidget()
            self.tab_widget.setVisible(False)
            self.pivot_tab = QWidget()
            self.filters_tab = QWidget()
            self.tab_widget.addTab(self.pivot_tab, "Pivot Parts")
            self.tab_widget.addTab(self.filters_tab, "Filters")

            #Pivot Detials Tab Widget
            pivot_layout = QVBoxLayout(self.pivot_tab)
            self.rows_list = self.create_list_widget("Rows")
            self.values_list = self.create_list_widget("Values")

            list_style = """
            QListWidget {
                border: 1px solid #d3d3d3;
                background-color: #ffffff;
            }

            QListWidget::item {
                padding: 2px 5px;
                border-bottom: 1px solid #f0f0f0;
            }

            QListWidget::item:hover {
                background-color: #f5f5f5;
            }

            QListWidget::item:selected {
                background-color: #e6f3ff;
                color: black;
            }
            """

            self.rows_list.setStyleSheet(list_style)
            self.values_list.setStyleSheet(list_style)

            pivot_layout.addWidget(self.rows_list)
            pivot_layout.addWidget(self.values_list)

            #Filters Tab Widget
            filters_layout = QVBoxLayout(self.filters_tab)
            self.filters_scroll = QScrollArea()
            self.filters_scroll.setWidgetResizable(True)
            self.filters_content = QWidget()
            self.filters_layout = QVBoxLayout(self.filters_content)
            self.filters_scroll.setWidget(self.filters_content)
            filters_layout.addWidget(self.filters_scroll)

            self.add_filter_button = QPushButton('+', self)
            self.add_filter_button.clicked.connect(self.add_filter)
            filters_layout.addWidget(self.add_filter_button)

            left_layout.addWidget(self.tab_widget)
            self.left_widget.setLayout(left_layout)

            #Right Side Widget
            right_widget = QWidget()
            right_layout = QVBoxLayout(right_widget)
            splitter = QSplitter(Qt.Orientation.Vertical)

            #Button Widgets
            button_layout = QHBoxLayout()
            file_button1 = QPushButton('Choose file for Current Period', self)
            file_button1.clicked.connect(lambda: self.load_csv(1))
            file_button2 = QPushButton('Choose file for Past Period', self)
            file_button2.clicked.connect(lambda: self.load_csv(2))
            generate_button = QPushButton('Generate Pivot', self)
            generate_button.clicked.connect(self.generate_pivot_table)
            export_button = QPushButton('Export', self)
            export_button.clicked.connect(self.exportXLSX)
            load_config_button = QPushButton('Load Configuration', self)
            load_config_button.clicked.connect(self.load_configuration)
            export_config_button = QPushButton('Export Configuration', self)
            export_config_button.clicked.connect(self.export_configuration)

            button_layout.addWidget(load_config_button)
            button_layout.addWidget(file_button1)
            button_layout.addWidget(file_button2)
            button_layout.addWidget(generate_button)
            button_layout.addWidget(export_config_button)
            button_layout.addWidget(export_button)

            self.next_config_button = QPushButton('Next Configuration', self)
            self.next_config_button.clicked.connect(self.load_next_configuration)
            self.next_config_button.setEnabled(False)
            button_layout.addWidget(self.next_config_button)

            right_layout.addLayout(button_layout)

            button_style = """
            QPushButton {
                padding: 10px;
                border: 2px solid #d9d9d9;
                border-radius: 0px;
                font-weight: bold;
                background-color: #f0f0f0;
                color: black;
            }
            QPushButton:pressed {
                background-color: #e0e0e0;
            }
            """

            file_button1.setStyleSheet(button_style)
            file_button2.setStyleSheet(button_style)
            generate_button.setStyleSheet(button_style)
            export_button.setStyleSheet(button_style)
            load_config_button.setStyleSheet(button_style)
            export_config_button.setStyleSheet(button_style)

            self.tree_tab_widget = QTabWidget()
            self.tree_tab_widget.setTabsClosable(True)
            self.tree_tab_widget.tabCloseRequested.connect(self.close_tab)
            self.tree_tab_widget.currentChanged.connect(self.update_current_tab_settings)

            self.add_tab_button = QToolButton(self)
            self.add_tab_button.setText('+')
            self.add_tab_button.clicked.connect(self.add_new_tab)
            self.tree_tab_widget.setCornerWidget(self.add_tab_button, Qt.Corner.TopRightCorner)

            self.add_new_pivot_tab()

            splitter.addWidget(self.tree_tab_widget)

            #Commentary Widget
            commentary_widget = QWidget()
            commentary_layout = QVBoxLayout(commentary_widget)
            template_label = QLabel("Commentary Template:")
            self.template_edit = QTextEdit()
            self.template_edit.setPlaceholderText("Enter Commentary Template Here...Components of Tree can be inserted using {Component.Variance} etc.")
            generate_commentary_button = QPushButton('Generate Commentary', self)
            generate_commentary_button.clicked.connect(self.generate_commentary)
            self.commentary_display = QTextEdit()
            self.commentary_display.setReadOnly(True)

            commentary_layout.addWidget(template_label)
            commentary_layout.addWidget(self.template_edit)
            commentary_layout.addWidget(generate_commentary_button)
            commentary_layout.addWidget(self.commentary_display)
            splitter.addWidget(commentary_widget)

            commentary_widget.setStyleSheet("""
                QWidget {
                    padding: 0px;
                    border: 0px;
                }
                QLabel {
                    font-weight: bold;
                    padding: 5px;
                    border: 2px solid #d9d9d9;
                    border-radius: 0px;
                    background-color: #f0f0f0;
                }
                QPushButton {
                    padding: 10px;
                    border: 2px solid #d9d9d9;
                    border-radius: 0px;
                    font-weight: bold;
                    background-color: #f0f0f0;
                    color: black;
                }
                QPushButton:pressed {
                    background-color: #e0e0e0;
                    border-style: inset;
                }
                QTextEdit {
                    border: 2px solid #d9d9d9;
                    border-radius: 0px;
                    padding: 5px;
                    background-color: #ffffff;
                }
            """)

            #Splitter Widget
            splitter.setSizes([900, 400])
            right_layout.addWidget(splitter)

            main_splitter = QSplitter(Qt.Orientation.Horizontal)
            main_splitter.addWidget(self.left_widget)
            main_splitter.addWidget(right_widget)
            main_splitter.setSizes([50, 1550])
            main_layout.addWidget(main_splitter)

            self.csv_path1 = ""
            self.csv_path2 = ""

    def load_configuration(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Configuration File", "", "Excel Files (*.xlsx)")
        if file_name:
            try:
                self.configurations = self.read_xlsx_configuration(file_name)
                if not self.configurations:
                    QMessageBox.warning(self, "Warning", "No configurations found in the file.")
                    return

                self.current_config_index = 0
                self.load_current_configuration()
                self.next_config_button.setEnabled(len(self.configurations) > 1)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load configuration: {str(e)}")

    def read_xlsx_configuration(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        configurations = []
        current_config = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            config_id = row[0]

            if config_id and config_id != 'X':
                if current_config:
                    configurations.append(current_config)
                current_config = {
                    'id': config_id,
                    'pivots': [],
                    'tables': [],
                    'manual_components': [],
                    'commentary_template': ''
                }
                current_config['pivot_count'] = row[1] or 0
                current_config['table_count'] = row[2] or 0
                current_config['manual_count'] = row[3] or 0

                pivot_id = row[4]
                if pivot_id:
                    current_config['pivots'].append({
                        'id': pivot_id,
                        'current_csv': self.resolve_path(file_path, row[5]),
                        'past_csv': self.resolve_path(file_path, row[6]),
                        'rows': row[7].split(',') if row[7] else [],
                        'values': row[8].split(',') if row[8] else [],
                        'filters': self.parse_filters(row[9])
                    })

                table_id = row[10]
                if table_id:
                    current_config['tables'].append({
                        'id': table_id,
                        'csv_path': self.resolve_path(file_path, row[11])
                    })

                manual_component = row[12]
                if manual_component:
                    current_config['manual_components'].append({
                        'id': manual_component,
                        'components': row[13]
                    })

                current_config['commentary_template'] = row[14] or ''
            elif config_id == 'X':
                if current_config is None:
                    raise ValueError("Configuration format error: 'x' found before any configuration started")

                pivot_id = row[4]
                if pivot_id:
                    current_config['pivots'].append({
                        'id': pivot_id,
                        'current_csv': self.resolve_path(file_path, row[5]),
                        'past_csv': self.resolve_path(file_path, row[6]),
                        'rows': row[7].split(',') if row[7] else [],
                        'values': row[8].split(',') if row[8] else [],
                        'filters': self.parse_filters(row[9])
                    })
                table_id = row[10]
                if table_id:
                    current_config['tables'].append({
                        'id': table_id,
                        'csv_path': self.resolve_path(file_path, row[11])
                    })


                manual_component_id = row[12]
                if manual_component_id:
                    current_config['manual_components'].append({
                      'id': manual_component_id,
                      'components': row[13],
                    })

            elif not any(row):
                if current_config:
                    configurations.append(current_config)
                    current_config = None
            else:
                raise ValueError(f"Unexpected row format in configuration file: {row}")

        if current_config:
            configurations.append(current_config)
        return configurations

    def resolve_path(self, config_file_path, csv_path):
        if not csv_path:
            return ''
        if os.path.isabs(csv_path):
            return csv_path
        return os.path.join(os.path.dirname(config_file_path), csv_path)

    def parse_filters(self, filters_str):
        if not filters_str:
            return {}
        filters = {}
        for filter_item in filters_str.split(';'):
            if ':' in filter_item:
                col, vals = filter_item.split(':')
                filters[col] = vals.split(',')
        return filters

    def load_current_configuration(self):
            if not self.configurations or self.current_config_index >= len(self.configurations):
                return

            config = self.configurations[self.current_config_index]

            while self.tree_tab_widget.count() > 0:
                self.tree_tab_widget.removeTab(0)
            self.tab_data.clear()

            for pivot in config['pivots']:
                self.load_pivot_configuration(pivot)

            for table in config['tables']:
                self.load_table_configuration(table)

            index = 1
            for manual in config['manual_components']:
                self.load_manual_configuration(manual,index)
                index += 1

            self.load_commentary_configuration(config['commentary_template'])

            # Generate pivot tables and commentary
            for i in range(self.tree_tab_widget.count()):
                if self.tab_data[i]['type'] == 'pivot':
                    self.tree_tab_widget.setCurrentIndex(i)
                    self.generate_pivot_table()

            self.generate_commentary()

    def load_next_configuration(self):
        if self.current_config_index < len(self.configurations) - 1:
            self.current_config_index += 1
            self.load_current_configuration()
        else:
            QMessageBox.information(self, "Info", "This is the last configuration.")

    def load_pivot_configuration(self, pivot_config):
        if pivot_config['current_csv'] == "" and pivot_config['past_csv'] == "":
          pass
        else:
          self.csv_path1 = pivot_config['current_csv']
          self.csv_path2 = pivot_config['past_csv']

        self.df1 = pd.read_csv(self.csv_path1)
        self.df2 = pd.read_csv(self.csv_path2)

        self.add_new_pivot_tab()
        current_index = self.tree_tab_widget.count() - 1
        self.tab_data[current_index]['rows'] = pivot_config['rows']
        self.tab_data[current_index]['values'] = pivot_config['values']
        self.tab_data[current_index]['filters'] = pivot_config['filters']
        self.df1_paths[current_index] = pivot_config['current_csv']
        self.df2_paths[current_index] = pivot_config['past_csv']

        self.update_current_tab_settings()

    def load_table_configuration(self, table_config):
        self.add_new_table_tab()
        current_index = self.tree_tab_widget.count() - 1
        table_widget = self.tab_data[current_index]['widget']
        self.load_table_file(table_widget, table_config['csv_path'])
        self.tab_data[current_index]['csv_path'] = table_config['csv_path']

    def load_manual_configuration(self, manual,index):
        prefix = f'Manual{index}_'
        print(manual)
        components_str = manual["components"]

        if components_str:
            self.add_new_manual_component_tab()
            current_index = self.tree_tab_widget.count() - 1
            tree_widget = self.tab_data[current_index]['widget']

            components = components_str.split(';')
            for component in components:
                name, formula = component.split(':')
                self.add_manual_component_to_tree(tree_widget, name, formula)

    def load_commentary_configuration(self, template):
        self.template_edit.setPlainText(template)

    def export_configuration(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Export Configuration File", "", "Excel Files (*.xlsx)")
        if file_name:
            try:
                if os.path.exists(file_name):
                    wb = openpyxl.load_workbook(file_name)
                    sheet = wb.active
                    startrow = sheet.max_row + 1
                else:
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    sheet.title = "Configurations"
                    headers = ["Configuration ID", "Number of Pivots", "Number of Tables", "Number of Manual Components",
                               "Pivot ID", "Current CSV Path", "Past CSV Path", "Rows", "Values", "Filters",
                               "Table ID", "Table CSV Path", "Manual Components", "Formula", "Commentary Template"]
                    for col, header in enumerate(headers, start=1):
                        sheet.cell(row=1, column=col, value=header)
                    startrow = 2

                config_id = f"{startrow-1}"
                pivot_count = sum(1 for tab in self.tab_data if tab['type'] == 'pivot')
                table_count = sum(1 for tab in self.tab_data if tab['type'] == 'table')
                manual_count = sum(1 for tab in self.tab_data if tab['type'] == 'manual_component')

                sheet.cell(row=startrow, column=1, value=config_id)
                sheet.cell(row=startrow, column=2, value=pivot_count)
                sheet.cell(row=startrow, column=3, value=table_count)
                sheet.cell(row=startrow, column=4, value=manual_count)

                maxrow = startrow
                rowp = startrow
                for index,tab in enumerate(self.tab_data):
                  if tab['type'] == 'pivot':
                      if rowp != startrow:
                        sheet.cell(row=rowp, column=1, value='X')
                      sheet.cell(row=rowp, column=5, value=f"Pivot{index+1}")
                      sheet.cell(row=rowp, column=6, value=self.df1_paths.get(index, ''))
                      sheet.cell(row=rowp, column=7, value=self.df2_paths.get(index, ''))
                      sheet.cell(row=rowp, column=8, value=','.join(tab['rows']))
                      sheet.cell(row=rowp, column=9, value=','.join(tab['values']))
                      sheet.cell(row=rowp, column=10, value=';'.join(f"{col}:{','.join(vals)}" for col, vals in tab['filters'].items()))
                      rowp += 1


                if rowp > maxrow : maxrow = rowp
                rowp = startrow
                for tab in self.tab_data:
                    if tab['type'] == 'table':
                        if rowp != startrow:
                          sheet.cell(row=rowp, column=1, value='X')
                        sheet.cell(row=rowp, column=11, value=f"Table{rowp}")
                        sheet.cell(row=rowp, column=12, value=tab.get('csv_path', ''))
                        rowp += 1

                if rowp > maxrow : maxrow = rowp
                rowp = startrow
                index = 1
                for tab in self.tab_data:
                    if tab['type'] == 'manual_component':
                        if rowp != startrow:
                            sheet.cell(row=rowp, column=1, value='X')

                        components = ""
                        for component in tab.get('components', []):
                            com = component["name"] + ":" + component["formula"] + ";"
                            components += com
                        components = components[:-1]
                        print(components)
                        sheet.cell(row=rowp, column=13, value=index)
                        sheet.cell(row=rowp, column=14, value=components)
                        index+=1
                        rowp += 1

                rowp = startrow
                sheet.cell(row=rowp, column=15, value=self.template_edit.toPlainText())

                sheet.cell(row=maxrow, column=1,value='X')
                for col in range(1, 16):
                    cell = sheet.cell(row=maxrow+1, column=col)
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

                for col in range(1, 16):
                    sheet.column_dimensions[get_column_letter(col)].width = 20

                wb.save(file_name)
                QMessageBox.information(self, "Success", f"Configuration exported to {file_name}")

            except Exception as e:
                 QMessageBox.critical(self, "Error", f"Failed to export configuration: {str(e)}")

    def toggle_navbar(self):
        current_width = self.left_widget.width()
        new_width = 300 if current_width == 50 else 50
        self.left_widget.setFixedWidth(new_width)
        self.tab_widget.setVisible(new_width == 300)
        self.toggle_nav_button.setText('▶' if new_width == 50 else '◀')

    def create_list_widget(self, title):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        label = QLabel(title)
        list_widget = QListWidget()
        list_widget.setSelectionMode(QListWidget.MultiSelection)
        layout.addWidget(label)
        layout.addWidget(list_widget)
        return widget

    def load_csv(self, date_num):
        file_name, _ = QFileDialog.getOpenFileName(self, f"Open CSV File for Date {date_num}", "", "CSV Files (*.csv)")
        if file_name:
            current_index = self.tree_tab_widget.currentIndex()
            if date_num == 1:
                self.df1 = pd.read_csv(file_name)
                df = self.df1
                df = df.convert_dtypes()
                df['B'] = pd.to_numeric(df['B'])
                self.df1 = df
                self.df1_paths[current_index] = file_name
                self.update_current_tab_settings()
            else:
                self.df2 = pd.read_csv(file_name)
                df = self.df2
                df = df.convert_dtypes()
                df['B'] = pd.to_numeric(df['B'])
                self.df2 = df
                self.df2_paths[current_index] = file_name
            print(f"Loaded CSV for Date {date_num}")

    def update_lists(self, df, tab_data):
        if df is not None:
            columns = df.columns.tolist()
            for list_widget, data_key in [(self.rows_list, 'rows'), (self.values_list, 'values')]:
                list_widget_obj = list_widget.findChild(QListWidget)
                if list_widget_obj is not None:
                    list_widget_obj.clear()
                    list_widget_obj.addItems(columns)
                    for i in range(list_widget_obj.count()):
                        item = list_widget_obj.item(i)
                        if item is not None:
                            item.setSelected(item.text() in tab_data[data_key])

            self.clear_filters()

    def clear_lists(self):
        for list_widget in [self.rows_list, self.values_list]:
            list_widget_obj = list_widget.findChild(QListWidget)
            list_widget_obj.clear()
        self.clear_filters()

    def clear_filters(self):
        while self.filters_layout.count():
            child = self.filters_layout.takeAt(0)
            if child is not None:
                widget = child.widget()
                if widget is not None:
                    widget.deleteLater()

    def display_existing_filters(self, filters):
        self.clear_filters()
        for column, selected_values in filters.items():
            self.add_filter_silent(column, selected_values)

    def add_filter_silent(self, column, selected_values):
        if self.df1 is not None:
            filter_widget = QWidget()
            filter_layout = QVBoxLayout(filter_widget)
            filter_label = QLabel(column)
            filter_layout.addWidget(filter_label)

            list_widget = QListWidget()
            list_widget.setSelectionMode(QListWidget.MultiSelection)
            list_widget.addItems(self.df1[column].unique().astype(str))

            if selected_values:
                for i in range(list_widget.count()):
                    item = list_widget.item(i)
                    if item is not None:
                        item.setSelected(item.text() in selected_values)

            filter_layout.addWidget(list_widget)
            self.filters_layout.addWidget(filter_widget)

            current_index = self.tree_tab_widget.currentIndex()
            self.tab_data[current_index]['filters'][column] = selected_values or []

            print(f"Added filter for column: {column}")

    def add_filter(self):
        if self.df1 is not None:
            column, ok = QInputDialog.getItem(self, "Select Column", "Choose a column to filter:", self.df1.columns.tolist(), 0, False)
            if ok and column:
                selected_values = self.tab_data[self.tree_tab_widget.currentIndex()]['filters'].get(column, [])

                filter_widget = QWidget()
                filter_layout = QVBoxLayout(filter_widget)
                filter_label = QLabel(column)
                filter_layout.addWidget(filter_label)

                list_widget = QListWidget()
                list_widget.setSelectionMode(QListWidget.MultiSelection)
                list_widget.addItems(self.df1[column].unique().astype(str))

                if selected_values:
                    for i in range(list_widget.count()):
                        item = list_widget.item(i)
                        if item is not None:
                            item.setSelected(item.text() in selected_values)

                filter_layout.addWidget(list_widget)
                self.filters_layout.addWidget(filter_widget)

                current_index = self.tree_tab_widget.currentIndex()
                self.tab_data[current_index]['filters'][column] = selected_values or []

                print(f"Added filter for column: {column}")

    def add_new_tab(self):
        tab_type, ok = QInputDialog.getItem(self, "Select Tab Type", "Choose the type of tab:",
                                            ["Pivot", "Table", "Manual Component"], 0, False)
        if ok:
            if tab_type == "Pivot":
                self.add_new_pivot_tab()
            elif tab_type == "Table":
                self.add_new_table_tab()
            else:
                self.add_new_manual_component_tab()

    def add_new_manual_component_tab(self):
            new_tree_widget = QTreeWidget()
            new_tree_widget.setHeaderLabels(['Component', 'Current', 'Past', 'Variance', 'Variance Change (%)', 'Status'])
            new_tree_widget.setAlternatingRowColors(True)
            new_tree_widget.setSortingEnabled(True)
            new_tree_widget.setAnimated(True)
            header = new_tree_widget.header()
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
            new_tree_widget.setStyleSheet("""
              QTreeWidget {
                  border: 1px solid #d3d3d3;
                  background-color: #ffffff;
              }

              QTreeWidget::item {
                  border-bottom: 1px solid #d3d3d3;
                  border-right: 1px solid #d3d3d3;
                  padding-left: 10px;
                  padding-top: 5px;
                  padding-bottom: 5px;
              }

              QTreeWidget::item:selected {
                  background-color: #e6e6e6;
              }

              QTreeWidget::item:hover {
                  background-color: #f0f0f0;
              }

              QHeaderView::section {
                  background-color: #800000;
                  color: white;
                  font-weight: bold;
                  border: 1px solid #d3d3d3;
                  padding-left: 10px;
                  padding-top: 5px;
                  padding-bottom: 5px;
              }
            """)

            add_component_button = QPushButton("Add Component")
            add_component_button.clicked.connect(lambda: self.add_manual_component(new_tree_widget))

            container = QWidget()
            layout = QVBoxLayout(container)
            layout.addWidget(add_component_button)
            layout.addWidget(new_tree_widget)

            new_tab_data = {
                'type': 'manual_component',
                'widget': new_tree_widget,
                'components': []
            }
            self.tab_data.append(new_tab_data)

            tab_index = self.tree_tab_widget.addTab(container, f"Manual {len(self.tab_data)}")
            self.tree_tab_widget.setCurrentIndex(tab_index)

            manual_name = f"Manual{self.tree_tab_widget.count()}"
            self.commentary_data[manual_name] = {}

    def add_new_table_tab(self):
        new_table_widget = QTableWidget()
        new_table_widget.setColumnCount(0)
        new_table_widget.setRowCount(0)

        stylesheet = """
        QTableWidget {
            border: 1px solid #d3d3d3;
            background-color: #ffffff;
            gridline-color: #d3d3d3;
        }

        QTableWidget::item {
            border-bottom: 1px solid #d3d3d3;
            border-right: 1px solid #d3d3d3;
            padding-left: 10px;
            padding-right: 10px;
            padding-top: 5px;
            padding-bottom: 5px;
        }

        QTableWidget::item:selected {
            background-color: #e6e6e6;
        }

        QTableWidget::item:hover {
            background-color: #f0f0f0;
        }

        QHeaderView::section {
            background-color: #800000;
            color: white;
            font-weight: bold;
            border: 1px solid #d3d3d3;
            padding-left: 10px;
            padding-right: 10px;
            padding-top: 5px;
            padding-bottom: 5px;
        }

        QTableCornerButton::section {
            background-color: #800000;
            border: 1px solid #d3d3d3;
        }
        """
        new_table_widget.setStyleSheet(stylesheet)

        browse_button = QPushButton("Browse")
        browse_button.clicked.connect(lambda: self.load_table_file(new_table_widget))

        container = QWidget()
        layout = QVBoxLayout(container)
        layout.addWidget(browse_button)
        layout.addWidget(new_table_widget)

        tab_index = self.tree_tab_widget.addTab(container, f"Table {self.tree_tab_widget.count() + 1}")
        self.tree_tab_widget.setCurrentIndex(tab_index)

        new_tab_data = {
            'type': 'table',
            'widget': new_table_widget,
            'data': None,
            'csv_path': None
        }
        self.tab_data.append(new_tab_data)

        #table_name = f"Table_{self.tree_tab_widget.count()}"
        #self.commentary_data[table_name] = {}

    def load_table_file(self, table_widget, file_name=None):
        if file_name is None:
            file_name, _ = QFileDialog.getOpenFileName(self, "Open CSV File", "", "CSV Files (*.csv)")
        if file_name:
            df = pd.read_csv(file_name)
            self.display_table(table_widget, df)

            current_index = self.tree_tab_widget.currentIndex()
            self.tab_data[current_index]['data'] = df
            self.tab_data[current_index]['csv_path'] = file_name

            table_name = f"Table{current_index + 1}"
            self.update_table_commentary_data(table_name, df)


    def display_table(self, table_widget, df):
        table_widget.setColumnCount(len(df.columns))
        table_widget.setRowCount(len(df))
        table_widget.setHorizontalHeaderLabels(df.columns)

        for i in range(len(df)):
            for j, column in enumerate(df.columns):
                table_widget.setItem(i, j, QTableWidgetItem(str(df.iloc[i, j])))

        table_widget.resizeColumnsToContents()

    def update_table_commentary_data(self, table_name, df):
        self.commentary_data[table_name] = {}
        for column in df.columns:
            self.commentary_data[table_name][column] = {}
            for index, value in enumerate(df[column]):
                self.commentary_data[table_name][column][index + 1] = value

    def update_table_tab(self, index):
        table_data = self.tab_data[index]
        if table_data['data'] is not None:
            self.display_table(table_data['widget'], table_data['data'])

    def add_new_pivot_tab(self):
        new_tree_widget = QTreeWidget()
        new_tree_widget.setHeaderLabels(['Component', 'Current', 'Past', 'Variance', 'Variance Change (%)', 'Status'])
        new_tree_widget.setAlternatingRowColors(True)
        new_tree_widget.setSortingEnabled(True)
        new_tree_widget.setAnimated(True)
        new_tree_widget.header().setSectionResizeMode(QHeaderView.ResizeToContents)
        new_tree_widget.setStyleSheet("""
          QTreeWidget {
              border: 1px solid #d3d3d3;
              background-color: #ffffff;
          }

          QTreeWidget::item {
              border-bottom: 1px solid #d3d3d3;
              border-right: 1px solid #d3d3d3;
              padding-left: 10px;
              padding-top: 5px;
              padding-bottom: 5px;
          }

          QTreeWidget::item:selected {
              background-color: #e6e6e6;
          }

          QTreeWidget::item:hover {
              background-color: #f0f0f0;
          }

          QHeaderView::section {
              background-color: #800000;
              color: white;
              font-weight: bold;
              border: 1px solid #d3d3d3;
              padding-left: 10px;
              padding-top: 5px;
              padding-bottom: 5px;
          }
        """)

        new_tab_data = {
            'type': 'pivot',
            'rows': [],
            'values': [],
            'columns': [],
            'filters': {},
            'pivot_table1': None,
            'pivot_table2': None,
            'manual_components': []
        }
        self.tab_data.append(new_tab_data)

        tab_index = self.tree_tab_widget.addTab(new_tree_widget, f"Pivot {len(self.tab_data)}")
        self.tree_tab_widget.setCurrentIndex(tab_index)

        self.update_current_tab_settings()

    def close_tab(self, index):
        if self.tree_tab_widget.count() > 1:
            self.tree_tab_widget.removeTab(index)
            del self.tab_data[index]

    def update_current_tab_settings(self):
        current_index = self.tree_tab_widget.currentIndex()
        if current_index >= 0 and current_index < len(self.tab_data):
            current_data = self.tab_data[current_index]
            if current_data['type'] == 'pivot':
                if self.df1 is not None:
                    self.update_lists(self.df1, current_data)
                    print(current_data['filters'])
                    if(current_data['filters'] != {}):
                        self.display_existing_filters(current_data['filters'])
            elif current_data['type'] == 'table':
                self.update_table_tab(current_index)
        else:
            self.clear_lists()

    def generate_pivot_table(self):
        if self.df1 is not None and self.df2 is not None:
            current_index = self.tree_tab_widget.currentIndex()
            current_data = self.tab_data[current_index]

            rows = [item.text() for item in self.rows_list.findChild(QListWidget).selectedItems()]
            values = [item.text() for item in self.values_list.findChild(QListWidget).selectedItems()]

            current_data['rows'] = rows
            current_data['values'] = values

            filters = {}
            for i in range(self.filters_layout.count()):
                filter_widget = self.filters_layout.itemAt(i).widget()
                if filter_widget is not None:
                    column_label = filter_widget.layout().itemAt(0).widget()
                    list_widget = filter_widget.layout().itemAt(1).widget()
                    if column_label is not None and list_widget is not None:
                        column = column_label.text()
                        selected_values = [item.text() for item in list_widget.selectedItems()]
                        if selected_values:
                            filters[column] = selected_values

            current_data['filters'] = filters

            mask1 = pd.Series([True] * len(self.df1))
            for col, filter_values in filters.items():
                mask1 &= self.df1[col].astype(str).isin(filter_values)
            df1_filtered = self.df1[mask1]

            mask2 = pd.Series([True] * len(self.df2))
            for col, filter_values in filters.items():
                mask2 &= self.df2[col].astype(str).isin(filter_values)
            df2_filtered = self.df2[mask2]

            current_data['pivot_table1'] = pd.pivot_table(df1_filtered, values=values, index=rows, aggfunc='sum')
            current_data['pivot_table2'] = pd.pivot_table(df2_filtered, values=values, index=rows, aggfunc='sum')

            self.update_tree_widget(rows, current_index)

    def update_tree_widget(self, rows, tab_index):
        tree_widget = self.tree_tab_widget.widget(tab_index)
        tree_widget.clear()

        current_data = self.tab_data[tab_index]
        pivot_table1 = current_data['pivot_table1']
        pivot_table2 = current_data['pivot_table2']

        if pivot_table1 is None or pivot_table2 is None:
            return

        pivot_name = f"Pivot{tab_index + 1}"
        self.commentary_data[pivot_name] = {}

        def add_children(parent, df1, df2, path=None):
            if path is None:
                path = []

            if len(path) >= len(rows):
                return

            children = []
            if len(rows) == 1:
                for value in df1.index:
                    child = QTreeWidgetItem(parent)
                    child.setText(0, f"{rows[0]}: {value}")

                    component_name = str(value)
                    value1 = df1.loc[value].sum()
                    value2 = df2.loc[value].sum() if value in df2.index else 0
                    self.set_comparison_values(child, value1, value2, component_name, pivot_name, path)
                    children.append(child)
            else:
                for value in df1.index.get_level_values(len(path)).unique():
                    child_df1 = df1.xs(value, level=len(path), drop_level=False)
                    child_df2 = df2.xs(value, level=len(path), drop_level=False) if value in df2.index.get_level_values(len(path)) else None

                    child = QTreeWidgetItem(parent)
                    child.setText(0, f"{rows[len(path)]}: {value}")

                    current_path = path + [value]
                    component_name = '.'.join(map(str, current_path))

                    if len(path) == len(rows) - 1:
                        value1 = child_df1.values.sum()
                        value2 = child_df2.values.sum() if child_df2 is not None else 0
                        self.set_comparison_values(child, value1, value2, component_name, pivot_name, current_path)
                    else:
                        value1 = child_df1.values.sum()
                        value2 = child_df2.values.sum() if child_df2 is not None else 0
                        self.set_comparison_values(child, value1, value2, component_name, pivot_name, current_path)
                        add_children(child, child_df1, child_df2 if child_df2 is not None else pd.DataFrame(), current_path)

                    children.append(child)

            children.sort(key=lambda x: float(x.text(4).rstrip('%')) if x.text(4) != "N/A" else 0, reverse=True)
            for i, child in enumerate(children):
                parent.insertChild(i, child)

        add_children(tree_widget.invisibleRootItem(), pivot_table1, pivot_table2)

        for name in current_data.get('manual_components', []):
            itemM = QTreeWidgetItem(tree_widget)
            itemM.setText(0, f"Manual: {name}")
            component_data = self.commentary_data[pivot_name].get(name, {})
            if component_data:
                self.set_comparison_values(itemM, component_data['value1'], component_data['value2'], name, pivot_name, [name])

        tree_widget.expandAll()

        for column in range(tree_widget.columnCount()):
            tree_widget.resizeColumnToContents(column)

    def set_comparison_values(self, item, value1, value2, component_name, pivot_name, path):
        value1 = 0 if pd.isna(value1) else value1
        value2 = 0 if pd.isna(value2) else value2

        item.setText(1, f"{value1:.2f}")
        item.setText(2, f"{value2:.2f}")
        variance = value2 - value1
        item.setText(3, f"{variance:.2f}")
        if value1 != 0:
            percentage_change = (variance / value1) * 100
            item.setText(4, f"{percentage_change:.2f}%")
        else:
            percentage_change = None
            item.setText(4, "N/A")
        if variance > 0:
            change_type = "Increase"
        elif variance < 0:
            change_type = "Decrease"
        else:
            change_type = "No Change"
        item.setText(5, change_type)

        component_data = {
            'value1': value1,
            'value2': value2,
            'variance': variance,
            'percentage_change': percentage_change,
            'change_type': change_type,
        }

        # Update the commentary_data structure
        current_level = self.commentary_data[pivot_name]
        for i, part in enumerate(path):
            if i == len(path) - 1:
                current_level[part] = component_data
            else:
                if part not in current_level:
                    current_level[part] = {}
                current_level = current_level[part]

        if len(path) > 1:
            parent_path = path[:-1]
            parent_level = self.commentary_data[pivot_name]
            for part in parent_path:
                parent_level = parent_level[part]
            if 'top_children' not in parent_level:
                parent_level['top_children'] = []
            parent_level['top_children'].append((path[-1], abs(variance)))
            parent_level['top_children'].sort(key=lambda x: x[1], reverse=True)
            parent_level['top_children'] = parent_level['top_children'][:3]

    def generate_commentary(self):
        template_str = self.template_edit.toPlainText()
        env = Environment()

        try:
            ast = env.parse(template_str)
            undefined_vars = meta.find_undeclared_variables(ast)


            context = self.commentary_data

            template = Template(template_str)
            commentary = template.render(context)

            self.commentary_display.setPlainText(commentary)
        except KeyError as e:
            self.commentary_display.setPlainText(f"Error: {e}. This key is not found in the data.")
        except Exception as e:
            self.commentary_display.setPlainText(f"Error in template formatting: {e}")

    def add_manual_component(self, tree_widget):
        name, ok = QInputDialog.getText(self, "Manual Component", "Enter component name:")
        if ok and name:
            formula, ok = QInputDialog.getText(self, "Manual Component", "Enter formula (e.g., Pivot1.A + Table2.B - Pivot3.C):")
            if ok and formula:
                self.add_manual_component_to_tree(tree_widget, name, formula)

    def add_manual_component_to_tree(self, tree_widget, name, formula):
        current_index = self.tree_tab_widget.currentIndex()
        itemM = QTreeWidgetItem(tree_widget)
        itemM.setText(0, name)

        value1 = value2 = None
        error_message = None

        try:
            if ',' in formula:
              values = formula.split(',')
              values = [float(value.strip()) for value in values]
              value1 = values[0]
              value2 = values[1]
            else:
              value1 = self.calculate_manual_component(formula, 1)
              value2 = self.calculate_manual_component(formula, 2)
            self.set_comparison_values(itemM, value1, value2, name, f"Manual{current_index + 1}", [name])
        except Exception as e:
            error_message = str(e)
            itemM.setText(1, "Error")
            itemM.setText(2, "Error")
            itemM.setText(3, error_message)

        manual_name = f"Manual{current_index + 1}"
        if value1 is not None and value2 is not None:
            self.commentary_data[manual_name][name] = {
                'value1': value1,
                'value2': value2,
                'variance': value2 - value1,
                'percentage_change': ((value2 - value1) / value1 * 100) if value1 != 0 else None,
                'change_type': 'Increase' if value2 > value1 else 'Decrease' if value2 < value1 else 'No Change'
            }
        else:
            self.commentary_data[manual_name][name] = {
                'value1': None,
                'value2': None,
                'variance': None,
                'percentage_change': None,
                'change_type': 'Error',
                'error_message': error_message
            }

        self.tab_data[current_index]['components'].append({'name': name, 'formula': formula})

        print(f"Added manual component: {name}")

    def calculate_manual_component(self, formula, date_num):
        components = formula.split()
        result = 0
        operation = '+'

        for component in components:
            if component in ['+', '-', '*', '/']:
                operation = component
            else:
                value = self.get_component_value(component, date_num)
                if operation == '+':
                    result += value
                elif operation == '-':
                    result -= value
                elif operation == '*':
                    result *= value
                elif operation == '/':
                    result /= value

        return result

    def get_component_value(self, component, date_num):
        parts = component.split('.')

        tab_reference = parts[0]
        component_name = parts[1:]
        tab_type, tab_number = tab_reference[:-1],tab_reference[-1]

        if tab_type not in ['Pivot', 'Table', 'Manual']:
            raise ValueError(f"Invalid tab type in component reference: {tab_type}")

        data_key = f"{tab_type}{tab_number}"

        if data_key not in self.commentary_data:
            raise ValueError(f"No data found for {data_key}")

        if tab_type == "Pivot" or tab_type == "Manual":
            component_data = self.commentary_data[data_key]
            for i in range(len(component_name)):
                component_data = component_data[component_name[i]]

            if f'value{date_num}' not in component_data:
                raise ValueError(f"Value for date {date_num} not found in component {component_name}")

            return component_data[f'value{date_num}']
        elif tab_type == "Table":
            component_data = self.commentary_data[data_key]

            component_name[-1] = int(component_name[-1])

            for i in range(len(component_name)):
                component_data = component_data[component_name[i]]

            return component_data
        return component_data[f'value{date_num}']

    def get_component_value2(self, component, date_num):
        print(component)
        tab_type, component_name = component.split('.')
        tab_index = int(tab_type[-1])
        tab_type = tab_type[:-1]
        tab_index = int(tab_index) - 1

        print(tab_type)
        print(tab_index)

        print(self.commentary_data)
        if tab_type == 'Pivot':
            pivot_data = self.tab_data[tab_index]['pivot_table1' if date_num == 1 else 'pivot_table2']
            print(pivot_data)
            return pivot_data.loc[component_name].values[0]
        elif tab_type == 'Table':
            table_data = self.tab_data[tab_index]['data']
            return table_data[component_name].sum()
        elif tab_type == 'Manual':
            manual_data = self.commentary_data[f"Manual{tab_index + 1}"]
            return manual_data[component_name][f'value{date_num}']
        else:
            raise ValueError(f"Invalid component reference: {component}")

    def exportXLSX(self):
        if any(tab['type'] in ['pivot', 'table', 'manual_component'] for tab in self.tab_data):
            file_name, _ = QFileDialog.getSaveFileName(self, "Save XLSX File", "", "XLSX Files (*.xlsx)")
            if file_name:
                with pd.ExcelWriter(file_name, engine='xlsxwriter') as writer:
                    commentary_data = []
                    for config_index, config in enumerate(self.configurations):
                        config_name = f"Configuration {config_index + 1}"
                        target_variable = None

                        # Recreate the commentary_data for this specific configuration
                        self.load_current_configuration()  # This should reset and populate commentary_data for the current config

                        # Find target variable (first manual component or first pivot component)
                        for component in config.get('manual_components', []):
                            target_variable = component['id']
                            break

                        if target_variable is None:
                            for pivot in config.get('pivots', []):
                                if pivot['rows']:
                                    target_variable = f"Pivot{pivot['id']}.{pivot['rows'][0]}"
                                    break

                        if target_variable:
                            #current_value = self.get_component_value(target_variable, 1)
                            #past_value = self.get_component_value(target_variable, 2)

                            current_value = 12
                            past_value = 24
                            # Generate commentary for this configuration
                            template_str = config.get('commentary_template', '')
                            env = Environment()
                            template = Template(template_str)
                            context = self.commentary_data  # Now using the configuration-specific commentary_data
                            commentary = template.render(context)

                            commentary_data.append([
                                config_index + 1,
                                config_name,
                                current_value,
                                past_value,
                                commentary
                            ])

                        # Move to the next configuration
                        self.current_config_index += 1
                        if self.current_config_index >= len(self.configurations):
                            self.current_config_index = 0

                    # Reset to the first configuration after exporting
                    self.current_config_index = 0
                    self.load_current_configuration()

                    # Export Commentary sheet
                    commentary_df = pd.DataFrame(commentary_data, columns=[
                        'S.No.', 'Configuration Name', 'Target Variable Current',
                        'Target Variable Past', 'Commentary'
                    ])
                    commentary_df.to_excel(writer, sheet_name='Commentary', index=False)

                    # Adjust column widths in Commentary sheet
                    worksheet = writer.sheets['Commentary']
                    worksheet.set_column('A:A', 10)  # S.No.
                    worksheet.set_column('B:B', 20)  # Configuration Name
                    worksheet.set_column('C:D', 25)  # Target Variable columns
                    worksheet.set_column('E:E', 100)  # Commentary

                print(f"Variance Table and Commentary Exported to {file_name}")
if __name__ == '__main__':
    app = QApplication(sys.argv)
    font = QFont("Univers Next for HSBC W02 Rg", 12)
    app.setFont(font)
    ex = VarianceCommentaryTool()
    ex.show()
    sys.exit(app.exec_())
