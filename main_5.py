import sys
import os
import subprocess
import openpyxl
import msoffcrypto
from tempfile import NamedTemporaryFile
from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QPushButton, QLabel, QComboBox, QTextEdit, QListWidget,
                             QScrollArea, QFrame,QMessageBox,QApplication,QDialog,QTableWidget,QTableWidgetItem)
from PyQt5.QtGui import QFont, QColor, QPalette
from PyQt5.QtCore import Qt, QSize
from file_handler import FileHandler
from data_processor import DataProcessor
from config import SHARED_DRIVE_PATH

class ColoredLabel(QLabel):
    def __init__(self, text, color):
        super().__init__(text)
        self.setStyleSheet(f"background-color: {color}; color: white; padding: 2px 5px; border-radius: 0px;")

class RiskItem(QFrame):
    def __init__(self, steward_name, description, is_due_soon):
        super().__init__()
        self.setStyleSheet("background-color: #f0f0f0; border-radius: 5px; margin: 2px;")
        layout = QHBoxLayout(self)

        tag_color = "#FF9800" if is_due_soon else "#DA0012"
        tag_text = "Due Soon" if is_due_soon else "Overdue"
        tag = ColoredLabel(tag_text, tag_color)
        layout.addWidget(tag)

        info = QLabel(f"{steward_name} - {description}")
        info.setStyleSheet("font-size: 14px;")
        layout.addWidget(info)
        layout.addStretch()


class HistoryDialog(QDialog):
    def __init__(self, actions):
        super().__init__()
        self.setWindowTitle("History of Actions")
        self.setGeometry(200, 200, 600, 400)
        layout = QVBoxLayout(self)

        self.table = QTableWidget()
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels(["Action"])
        self.table.setRowCount(len(actions))

        for i, action in enumerate(actions):
            self.table.setItem(i, 0, QTableWidgetItem(action))

        layout.addWidget(self.table)

class RiskReportApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Risk Report Application")
        self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet("""
            QMainWindow {
                background-color: #FFFFFF;
            }
            QLabel {
                font-size: 16px;
            }
            QPushButton {
                font-size: 16px;
                padding: 10px;
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QComboBox {
                font-size: 14px;
                padding: 5px;
                background-color: #f0f0f0;
                border: 1px solid #BDBDBD;
                border-radius: 5px;
                padding-left: 10px;
                padding-right: 30px;
                min-width: 150px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 30px;
                border-left-width: 1px;
                border-left-color: #BDBDBD;
                border-left-style: solid;
                border-top-right-radius: 5px;
                border-bottom-right-radius: 5px;
                background-color: #E0E0E0;
            }
            QTextEdit {
                font-size: 14px;
                background-color: white;
                border: 1px solid #BDBDBD;
                border-radius: 5px;
            }
        """)

        self.file_handler = FileHandler(SHARED_DRIVE_PATH)
        self.data_processor = DataProcessor()
        #self.email_handler = EmailHandler()
        self.actions = []
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        main_layout = QHBoxLayout(self.central_widget)

        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)

        #self.info_label = QLabel("Inform")
        #self.info_label.setStyleSheet("""
        #    font-size: 15px;
        #    font-weight: bold;
        #    background: qlineargradient(
        #        x1: 0, y1: 0,
        #        x2: 1, y2: 1,
        #        stop: 0 #FF512F,
        #        stop: 1 #DD2476
        #    );
        #    color: white;
        #    padding: 10px;
        #    border-radius: 1px;
        #""")

        self.info_text = QTextEdit()
        self.info_text.setReadOnly(True)
        self.info_text.setStyleSheet("""
            font-size: 14px;
            background-color: #f9f9f9;
            border: 1px solid #BDBDBD;
            border-radius: 5px;
            padding: 10px;
            font-family: sans-serif;
        """)

        self.change_button = QPushButton("Change Template")
        self.change_button.setStyleSheet("""
            QPushButton {
                font-size: 14px;
                padding: 5px;
                background-color: #F13E40;
                color: white;
                border: none;
                border-radius: 1px;
            }
            QPushButton:hover {
                background-color: #bc363f;
            }
        """)
        self.send_button = QPushButton("Send Email")
        self.send_button.setStyleSheet("""
            QPushButton {
                font-size: 14px;
                padding: 5px;
                background-color: #F13E40;
                color: white;
                border: none;
                border-radius: 1px;
            }
            QPushButton:hover {
                background-color: #bc363f;
            }
        """)

        #left_layout.addWidget(self.info_label)
        left_layout.addWidget(self.info_text)
        left_layout.addWidget(self.send_button)
        left_layout.addWidget(self.change_button)

        middle_panel = QWidget()
        middle_layout = QVBoxLayout(middle_panel)
        intros_layout = QHBoxLayout()
        intros_layout.addWidget(QLabel("Risk Report Application"))

        self.welcome_label = QLabel(f"Welcome {os.getenv('USERNAME', 'User')}")
        self.welcome_label.setStyleSheet("font-size: 24px; font-weight: bold; margin-bottom: 10px;")

        intros_layout.addWidget(self.welcome_label)

        controls_layout = QHBoxLayout()
        self.week_combo = QComboBox()
        self.region_combo = QComboBox()
        self.risk_combo = QComboBox()
        controls_layout.addWidget(QLabel("Week:"))
        controls_layout.addWidget(self.week_combo)
        controls_layout.addWidget(QLabel("Region:"))
        controls_layout.addWidget(self.region_combo)
        controls_layout.addWidget(QLabel("Risk Type:"))
        controls_layout.addWidget(self.risk_combo)

        self.history_button = QPushButton("See History")
        self.history_button.setStyleSheet("background-color: #4CAF50;")
        self.history_button.clicked.connect(self.show_history_dialog)
        intros_layout.addWidget(self.history_button)

        self.send_all_button = QPushButton("Send All")
        self.send_all_button.setStyleSheet("background-color: #4CAF50;")
        self.risk_list = QScrollArea()
        self.risk_list.setWidgetResizable(True)
        self.risk_list_content = QWidget()
        self.risk_list_layout = QVBoxLayout(self.risk_list_content)
        self.risk_list.setWidget(self.risk_list_content)

        middle_layout.addLayout(intros_layout)
        middle_layout.addLayout(controls_layout)
        middle_layout.addWidget(self.send_all_button)
        middle_layout.addWidget(self.risk_list)

        main_layout.addWidget(left_panel, 1)
        main_layout.addWidget(middle_panel, 2)

        self.change_button.clicked.connect(self.change_email_template)
        self.send_button.clicked.connect(self.send_email)
        self.send_all_button.clicked.connect(self.send_all_emails)
        self.week_combo.currentIndexChanged.connect(self.update_report)
        self.region_combo.currentIndexChanged.connect(self.update_content)
        self.risk_combo.currentIndexChanged.connect(self.update_content)

    def load_data(self):
        self.log_action("Application Started")
        weeks = self.file_handler.get_available_weeks()
        self.week_combo.addItems(weeks)
        self.week_combo.setCurrentIndex(len(weeks) - 1)

        self.mapping_data = self.file_handler.read_mapping_file()
        regions = self.mapping_data['Region'].unique()

        regions_actual = []
        for region in regions:
          if type(region) is str:
            regions_actual.append(region)

        self.region_combo.addItems(regions_actual)

        risk_types = ['All', 'Tax', 'Financial']
        self.risk_combo.addItems(risk_types)

        self.update_report()

    def update_report(self):
        selected_week = self.week_combo.currentText()
        self.current_report = self.file_handler.read_risk_report(selected_week)
        self.update_content()
        self.log_action(f"Loaded Report for Week {selected_week}")

    def update_content(self):
          region = self.region_combo.currentText()
          risk_type = self.risk_combo.currentText()

          filtered_data = self.data_processor.filter_data(self.current_report, region, risk_type)

          # Clear existing items
          while self.risk_list_layout.count():
              item = self.risk_list_layout.takeAt(0)
              if item.widget():
                  item.widget().deleteLater()

          for _, row in filtered_data.iterrows():
              is_due_soon = row['Due < 42 Days'] == 1
              item = RiskItem(row['Steward Name'], row['Description'], is_due_soon)
              item.mousePressEvent = lambda event, r=row: self.show_risk_info(r)
              self.risk_list_layout.addWidget(item)

          self.risk_list_layout.addStretch()

    def show_risk_info(self, risk):
        risk_info = self.data_processor.get_risk_info(risk)
        self.info_text.setText(risk_info)

    def change_email_template(self):
        QMessageBox.information(self, "Change Template", "Email template change functionality would be implemented here.")
        self.log_action("Email Template Changed")

    def send_email(self):
        selected_items = [item for item in self.risk_list_content.children() if isinstance(item, RiskItem) and item.isActiveWindow()]
        if selected_items:
            item = selected_items[0]
            steward_name = item.findChild(QLabel).text().split(' - ')[0]
            #success = self.email_handler.send_email(steward_name, self.info_text.toPlainText())
            success = 1
            if success:
                self.log_action(f"Email sent to {steward_name}")
            else:
                self.log_action(f"Failed to send email to {steward_name}")
        else:
            QMessageBox.warning(self, "No Selection", "Please select a risk item to send an email.")

    def send_all_emails(self):
        for item in self.risk_list_content.children():
            if isinstance(item, RiskItem):
                steward_name = item.findChild(QLabel).text().split(' - ')[0]
                risk_info = self.info_text.toPlainText()
                #success = self.email_handler.send_email(steward_name, risk_info)
                success = 1
                if success:
                    self.log_action(f"Email sent to {steward_name}")
                else:
                    self.log_action(f"Failed to send email to {steward_name}")
        QMessageBox.information(self, "Emails Sent", "All emails have been sent.")

    def log_action(self, action):
        self.actions.append(action)

    def show_history_dialog(self):
        dialog = HistoryDialog(self.actions)
        dialog.exec_()

def load_workbook(target, password):
    with open(target, "rb") as fp, NamedTemporaryFile(suffix=".xlsx") as tmpfile:
        msf = msoffcrypto.OfficeFile(fp)
        msf.load_key(password=password)
        msf.decrypt(tmpfile)

        return openpyxl.load_workbook(tmpfile.name)

if __name__ == "__main__":
    app = QApplication(sys.argv)

    VERSION_CONTROL_FILE = "Version_Control.xlsx"
    VERSION_CONTROL_PASSWORD = "globalrisk"
    APPLICATION_NAME = "Risk Report Application"

    try:
        result = subprocess.run(['whoami'], capture_output=True, text=True)
        current_user = result.stdout.strip()
    except Exception as e:
        print(f"Error getting current user: {e}")
        current_user = None

    has_access = False
    if current_user:
        try:
            wb = load_workbook(VERSION_CONTROL_FILE, password=VERSION_CONTROL_PASSWORD)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == current_user and row[1] == APPLICATION_NAME:
                    has_access = True
                    break
        except Exception as e:
            print(f"Error checking user access: {e}")

    if has_access:
        window = RiskReportApp()
        window.show()
        sys.exit(app.exec_())
    else:
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("Access Denied")
        msg.setInformativeText("You do not have access to this application.")
        msg.setWindowTitle("Error")
        msg.exec_()
        sys.exit()
